import requests
import openpyxl
import os

# Steam API Information
API_KEY = os.getenv('STEAM_API_KEY', 'E38C84A35DD1546893005B2E960CA288')  # Use environment variable
STEAM_ID = '76561198074846013'

# URL for the Steam API GetOwnedGames method
url = f"https://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/?key={API_KEY}&steamid={STEAM_ID}&include_played_free_games=true&include_appinfo=false"

# Fetching owned games from Steam API
response = requests.get(url)
if response.status_code == 200:
    games_data = response.json().get('response', {}).get('games', [])
else:
    games_data = []

# Create a dictionary for easy access to playtime by App ID
playtime_dict = {
    game.get('appid', 'Unknown'): {
        'game_name': game.get('name', 'Unknown'),
        'playtime': round(game.get('playtime_forever', 0) / 60, 2)
    }
    for game in games_data
}

# Path to the existing spreadsheet
spreadsheet_path = r'D:\SteamHours\Steam Game V2.xlsx'

def get_total_games_and_hours(sheet):
    """
    Function to calculate the total number of games and total playtime in hours.
    
    Args:
        sheet: The sheet object containing the game data.
        
    Returns:
        A tuple containing (total_games, total_hours).
    """
    total_games = 0
    total_hours = 0.0
    
    # Iterate over the rows to count the games and sum the playtime (column C)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        app_id = row[0]
        playtime = row[2]  # Column C: Hours
        
        # Ensure that playtime is a valid number before adding
        if app_id and playtime is not None:
            try:
                # Attempt to convert playtime to float if it's not already
                playtime = float(playtime)
                total_games += 1
                total_hours += playtime
            except ValueError:
                pass
    total_hours = round(total_hours, 2)
    return total_games, total_hours

def get_average_playtime(sheet):
    """
    Function to calculate the average playtime per game.
    
    Args:
        sheet: The sheet object containing the game data.
        
    Returns:
        Average playtime per game as a float.
    """
    total_games, total_hours = get_total_games_and_hours(sheet)
    return total_hours / total_games if total_games > 0 else 0

def update_spreadsheet():
    """ 
    Function to update the spreadsheet with the latest game data. 
    """
    try:
        workbook = openpyxl.load_workbook(spreadsheet_path)
        if 'Hours' in workbook.sheetnames:
            sheet = workbook['Hours']

            # Update hours in the sheet (implementation here...)

            # Save the workbook
            workbook.save(spreadsheet_path)
    except Exception as e:
        print(f"Error updating spreadsheet: {e}")

# Optional: Only run this code if executing directly
if __name__ == '__main__':
    update_spreadsheet()
