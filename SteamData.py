import openpyxl
import logging
from datetime import datetime
import os

logging.basicConfig(level=logging.INFO)

def log_game_error(game_name, app_id, playtime_value, error_message):
    """Log game errors to a file instead of printing to terminal."""
    error_file_path = r'D:\SteamHours\game_errors.log'
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        with open(error_file_path, 'a', encoding='utf-8') as error_file:
            error_file.write(f"[{timestamp}] Game: {game_name} (App ID: {app_id}) - Playtime: {playtime_value} - Error: {error_message}\n")
    except Exception as e:
        print(f"Failed to write to error log: {e}")

def get_data_from_spreadsheet(spreadsheet_path=r'D:\SteamHours\ExcelFiles\steam_games_playtime.xlsx'):
    """Get total games and total hours from the spreadsheet."""
    try:
        workbook = openpyxl.load_workbook(spreadsheet_path)
        if 'Steam Games Playtime' in workbook.sheetnames:
            sheet = workbook['Steam Games Playtime']
            total_games = 0
            total_hours = 0.0

            # Iterate over the rows to count the games and sum the playtime (column C)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                game_name = row[0] if len(row) > 0 else "Unknown Game"  # Column A: Game Name
                app_id = row[1] if len(row) > 1 else None  # Column B: App ID
                playtime = row[2] if len(row) > 2 else None  # Column C: Hours

                if playtime is not None:
                    try:
                        playtime = float(playtime)
                        total_games += 1
                        total_hours += playtime
                    except ValueError:
                        # Log to file instead of printing to terminal
                        log_game_error(game_name, app_id, playtime, "Invalid playtime value - could not convert to float")
                        continue

            if total_games == 0:
                logging.info("No games found in the spreadsheet.")
                return 0, 0.0, 0.0

            average_playtime = round(total_hours / total_games, 2)
            return total_games, round(total_hours, 2), average_playtime
        else:
            logging.error("Sheet 'Steam Games Playtime' does not exist.")
            return 0, 0.0, 0.0
    except FileNotFoundError:
        logging.error(f"Spreadsheet not found at path: {spreadsheet_path}")
        return 0, 0.0, 0.0
    except openpyxl.utils.exceptions.InvalidFileException:
        logging.error(f"Invalid spreadsheet file: {spreadsheet_path}")
        return 0, 0.0, 0.0
    except Exception as e:
        logging.error(f"Unexpected error reading spreadsheet: {e}")
        return 0, 0.0, 0.0