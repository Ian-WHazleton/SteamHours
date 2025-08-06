import random
from PyQt6.QtWidgets import QLineEdit
import csv
import re
import json
import os
from datetime import datetime
from PyQt6.QtWidgets import QApplication, QMainWindow, QLabel, QHBoxLayout, QVBoxLayout, QWidget, QSpacerItem, QSizePolicy, QPushButton, QGridLayout, QMessageBox, QInputDialog, QCheckBox, QDialog, QFileDialog, QTextEdit
from custom_textbox import CustomTextBox
from PyQt6.QtGui import QFont
from PyQt6.QtCore import Qt, QTimer
import sys
import openpyxl
from SteamAPI_Caller import update_spreadsheet
from SteamData import get_data_from_spreadsheet
from steam_csv_importer import SteamCSVImporter
from game_search import calculate_similarity_score

class GameLookupDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Game Hours Lookup')
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #ccc;
                padding: 5px;
                margin: 5px 0;
            }
            QCheckBox {
                color: white;
                margin: 5px 0;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                background-color: white;
                border: 1px solid #ccc;
            }
            QCheckBox::indicator:checked {
                background-color: #4CAF50;
                border: 1px solid #4CAF50;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # Create layout
        layout = QVBoxLayout()
        
        # Add input field
        self.label = QLabel('Enter Steam App ID:')
        layout.addWidget(self.label)
        
        self.input_field = CustomTextBox()
        layout.addWidget(self.input_field)
        
        # Add checkbox
        self.api_checkbox = QCheckBox('Use Steam API (instead of spreadsheet)')
        layout.addWidget(self.api_checkbox)
        
        # Add buttons
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton('OK')
        self.cancel_button = QPushButton('Cancel')
        
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # Set focus on input field
        self.input_field.setFocus()
    
    def get_values(self):
        return self.input_field.text(), self.api_checkbox.isChecked()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # Add user attribute for Steam ID (load early)
        self.current_steam_id = self.load_user_preferences()  # Load from config file
        
        # Ensure user directory exists on startup
        self.ensure_user_directory()
        
        self.setWindowTitle(f"Steam Data Tracker - User: {self.current_steam_id}")

        # Set the main window background color
        self.setStyleSheet("background-color: #202020;")  # Dark gray background

        # Make the window cover the entire screen but remain windowed
        screen = QApplication.primaryScreen()
        screen_geometry = screen.geometry()
        self.setGeometry(screen_geometry)  # Set the window geometry to the screen size

        # Create a central widget
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        # Create a vertical layout for the boxes
        main_layout = QVBoxLayout()

        # Create a horizontal layout for the boxes
        boxes_layout = QHBoxLayout()
        boxes_layout.setSpacing(40)  # Set the gap between boxes (in pixels)

        # Create the first box (number and caption)
        self.total_games_label = self.add_number_box(boxes_layout, "0", "Total Games", "#FFCCCB")  # Light red
        
        # Create the second box (number and caption)
        self.total_hours_label = self.add_number_box(boxes_layout, "0.00", "Total Hours", "#ADD8E6")  # Light blue
        
        # Create the third box (number and caption)
        self.average_playtime_label = self.add_number_box(boxes_layout, "0.00", "Average Playtime", "#90EE90")  # Light green

        # Add the boxes layout to the main layout
        main_layout.addLayout(boxes_layout)

        # Define button names in a dictionary
        self.button_names = {
            1: "Update Steam Info",
            2: "Random Game",
            3: "Game Hours Lookup",
            4: "Import Costs from CSV",
            5: "Search Game Stats",
            6: "Button 6",
            7: "Button 7",
            8: "Button 8",
            9: "Button 9",
            10: "Button 10",
            11: "Button 11",
            12: "Change User"
        }

        # Create a grid layout for buttons
        button_grid = QGridLayout()

        # Add buttons to the grid layout using the names from the dictionary

        for i in range(1, 13):  # 12 buttons
            button = QPushButton(self.button_names[i])
            button.setMinimumWidth(400)
            button.setMaximumWidth(400)
            button.setMinimumHeight(100)
            button.setMaximumHeight(100)
            button.setStyleSheet("background-color: #D3D3D3;")

            # Connect specific actions to specific buttons
            if i == 1:
                button.clicked.connect(self.update_steam_data)
            elif i == 2:
                button.clicked.connect(self.select_random_game)
            elif i == 3:
                button.clicked.connect(self.lookup_game_hours)
            elif i == 4:
                button.clicked.connect(self.import_costs_from_csv)
            elif i == 5:
                button.clicked.connect(self.search_game_stats)
            elif i == 12:
                button.clicked.connect(self.change_user)

            row = (i - 1) // 3
            col = (i - 1) % 3
            button_grid.addWidget(button, row, col)

        # Add the button grid to the main layout
        main_layout.addLayout(button_grid)

        # Set the layout for the central widget
        central_widget.setLayout(main_layout)

        # Check for spreadsheet existence on startup
        import os
        spreadsheet_path = self.get_user_spreadsheet_path()

        if not os.path.exists(spreadsheet_path):
            # Ensure user directory exists
            self.ensure_user_directory()
            
            # Custom popup: Create empty sheet or close
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Missing User Spreadsheet")
            msg_box.setText(f"The spreadsheet for user {self.current_steam_id} was not found at:\n{spreadsheet_path}\n\nWould you like to create an empty sheet for this user?")
            msg_box.setIcon(QMessageBox.Icon.Critical)
            create_button = msg_box.addButton("Create Empty Sheet", QMessageBox.ButtonRole.AcceptRole)
            close_button = msg_box.addButton("Close App", QMessageBox.ButtonRole.RejectRole)
            msg_box.setStyleSheet("""
                QMessageBox {
                    background-color: #2b2b2b;
                    color: white;
                }
                QMessageBox QLabel {
                    color: white;
                }
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    margin: 2px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            msg_box.exec()

            # Check which button was pressed using buttonRole
            clicked_role = msg_box.buttonRole(msg_box.clickedButton())
            if clicked_role == QMessageBox.ButtonRole.AcceptRole:
                from SteamAPI_Caller import create_blank_steam_spreadsheet
                create_blank_steam_spreadsheet(spreadsheet_path)
            else:
                sys.exit()

        # Load initial data from spreadsheet on startup
        self.load_initial_data()

    def change_user(self):
        """Prompt for a new Steam ID and update the user."""
        # Create a styled input dialog
        dialog = QInputDialog(self)
        dialog.setWindowTitle("Change Steam User")
        dialog.setLabelText(f"Current Steam ID: {self.current_steam_id}\n\nEnter new Steam ID:")
        dialog.setTextValue(self.current_steam_id)
        dialog.setStyleSheet("""
            QInputDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit {
                background-color: #404040;
                color: white;
                border: 1px solid #666;
                padding: 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            steam_id = dialog.textValue().strip()
            if steam_id and steam_id != self.current_steam_id:
                # Basic validation - Steam IDs are 17-digit numbers
                if steam_id.isdigit() and len(steam_id) == 17:
                    old_id = self.current_steam_id
                    self.current_steam_id = steam_id
                    self.save_user_preferences()  # Save to config file
                    
                    # Ensure user directory exists
                    self.ensure_user_directory()
                    
                    # Reload data for new user
                    self.load_initial_data()
                    
                    # Update window title to show current user
                    self.setWindowTitle(f"Steam Data Tracker - User: {self.current_steam_id}")
                    
                    self.show_success_notification("User Changed", 
                        f"Steam ID changed from:\n{old_id}\nto:\n{self.current_steam_id}\n\nData loaded for new user.")
                else:
                    self.show_styled_message_box("Invalid Steam ID", 
                        "Steam ID must be a 17-digit number.\n\nExample: 76561198074846013", 
                        QMessageBox.Icon.Warning)
            elif not steam_id:
                self.show_styled_message_box("Change Cancelled", 
                    "No Steam ID entered.", QMessageBox.Icon.Information)
            else:
                self.show_styled_message_box("No Change", 
                    "Steam ID unchanged.", QMessageBox.Icon.Information)

    def load_user_preferences(self):
        """Load user preferences from config file."""
        config_path = 'config/user_preferences.json'
        default_steam_id = '76561198074846013'
        
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    return config.get('steam_id', default_steam_id)
        except Exception as e:
            print(f"Error loading user preferences: {e}")
        
        return default_steam_id
    
    def save_user_preferences(self):
        """Save user preferences to config file."""
        config_path = 'config/user_preferences.json'
        
        try:
            # Create config directory if it doesn't exist
            os.makedirs(os.path.dirname(config_path), exist_ok=True)
            
            config = {
                'steam_id': self.current_steam_id,
                'last_updated': datetime.now().isoformat()  # Proper timestamp
            }
            
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
                
        except Exception as e:
            print(f"Error saving user preferences: {e}")
            self.show_styled_message_box("Save Error", 
                f"Could not save user preferences: {str(e)}", 
                QMessageBox.Icon.Warning)

    def get_user_spreadsheet_path(self):
        """Get the spreadsheet path for the current user."""
        return f'ExcelFiles/{self.current_steam_id}/steam_games_playtime.xlsx'
    
    def ensure_user_directory(self):
        """Ensure the user's directory exists."""
        user_dir = f'ExcelFiles/{self.current_steam_id}'
        try:
            os.makedirs(user_dir, exist_ok=True)
            print(f"User directory ensured: {user_dir}")
            return user_dir
        except Exception as e:
            print(f"Error creating user directory {user_dir}: {e}")
            self.show_styled_message_box("Directory Error", 
                f"Could not create user directory: {str(e)}", 
                QMessageBox.Icon.Warning)
            return None

    def load_initial_data(self):
        """Load initial data from the spreadsheet on startup."""
        try:
            # Ensure user directory exists before trying to access spreadsheet
            self.ensure_user_directory()
            spreadsheet_path = self.get_user_spreadsheet_path()
            total_games, total_hours, average_playtime = get_data_from_spreadsheet(spreadsheet_path)
            
            # Update the labels with the loaded data
            self.total_games_label.setText(str(total_games))
            self.total_hours_label.setText(f"{total_hours:.2f}")
            self.average_playtime_label.setText(f"{average_playtime:.2f}")
        except Exception as e:
            print(f"Error loading initial data: {e}")
            # Keep default values (0, 0.00, 0.00) if there's an error

    def add_number_box(self, layout, number, caption, color):
        """Helper function to add a number box to the layout."""
        # Create a vertical layout for each box
        box_layout = QVBoxLayout()

        number_label = QLabel(number, self)
        number_label.setFont(QFont("Arial", 24))
        number_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        caption_label = QLabel(caption, self)
        caption_label.setFont(QFont("Arial", 12))
        caption_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Add the labels to the box layout
        box_layout.addWidget(number_label)
        box_layout.addWidget(caption_label)

        # Create a widget to hold the box layout and set its background color
        box_widget = QWidget(self)
        box_widget.setLayout(box_layout)
        box_widget.setStyleSheet(f"background-color: {color}; padding: 10px; border-radius: 5px;")

        # Set a fixed width for the box widget
        box_widget.setMinimumWidth(300)
        box_widget.setMaximumWidth(300)
        box_widget.setMinimumHeight(300)
        box_widget.setMaximumHeight(300)

        # Add the box widget to the horizontal layout
        layout.addWidget(box_widget)

        # Return the QLabel to update later
        return number_label

    def update_steam_data(self):
        """Update Steam data and show loading animation."""
        # Create a local update button
        self.update_button = QPushButton("Update")  # Assign to self

        # Disable the update button and show a throbber
        self.update_button.setText("Updating...")
        self.update_button.setEnabled(False)

        # Start the throbber
        self.start_throbber()

        # Update the spreadsheet in a separate thread
        QTimer.singleShot(100, self.perform_update)  # Simulate an async operation

    def perform_update(self):
        """Perform the update of Steam data."""
        try:
            # Update the spreadsheet
            from SteamAPI_Caller import get_api_key
            api_key = get_api_key()
            if not api_key:
                self.show_styled_message_box("Steam API Key Error", "Steam API key not found. Please check your .env file.", QMessageBox.Icon.Critical)
                return

            # Ensure user directory exists before accessing spreadsheet
            self.ensure_user_directory()
            spreadsheet_path = self.get_user_spreadsheet_path()
            
            # Pass user-specific Steam ID and spreadsheet path
            update_spreadsheet(steam_id=self.current_steam_id, spreadsheet_path=spreadsheet_path)
            
            # Get the updated values using SteamData
            total_games, total_hours, average_playtime = get_data_from_spreadsheet(spreadsheet_path)

            # Update the labels with the new data
            self.total_games_label.setText(str(total_games))
            self.total_hours_label.setText(f"{total_hours:.2f}")
            self.average_playtime_label.setText(f"{average_playtime:.2f}")

            # Show success notification
            message = f"Steam data updated successfully!\n\n"
            message += f"Total games: {total_games}\n"
            message += f"Total hours: {total_hours:.2f}"
            
            self.show_success_notification("Steam Data Updated!", message)
            
        except Exception as e:
            # Show error notification if update fails
            self.show_styled_message_box("Update Failed", f"Failed to update Steam data: {str(e)}", QMessageBox.Icon.Critical)
        finally:
            # Restore button state
            self.update_button.setText("Update Steam Info")

            # Stop the throbber
            self.stop_throbber()

    def start_throbber(self):
        """Start a simple throbber animation."""
        self.throbber_timer = QTimer(self)
        self.throbber_timer.timeout.connect(self.update_throbber)
        self.throbber_counter = 0
        self.throbber_text = "Updating"  # Base text for the throbber
        self.throbber_timer.start(500)  # Update every 500ms

    def update_throbber(self):
        """Update the throbber text."""
        # Add a dot for the throbber effect
        self.throbber_counter = (self.throbber_counter + 1) % 4
        dots = "." * self.throbber_counter
        self.update_button.setText(self.throbber_text + dots)

    def stop_throbber(self):
        """Stop the throbber timer."""
        self.throbber_timer.stop()

    def select_random_game(self):
        """Select a random game from the spreadsheet and display its name."""
        try:
            # Ensure user directory exists before accessing spreadsheet
            self.ensure_user_directory()
            spreadsheet_path = self.get_user_spreadsheet_path()
            workbook = openpyxl.load_workbook(spreadsheet_path)

            if 'Steam Games Playtime' in workbook.sheetnames:
                sheet = workbook['Steam Games Playtime']
                games = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    app_id, game_name = row[1], row[0]  # Note: swapped order - Game Name is column A, App ID is column B
                    if game_name:
                        games.append(game_name)

                if games:
                    random_game = random.choice(games)
                    # Display the selected game name in a pop-up
                    self.show_random_game_popup(random_game)
        except FileNotFoundError:
            self.show_styled_message_box("File Error", "Spreadsheet file not found.", QMessageBox.Icon.Warning)
        except Exception as e:
            self.show_styled_message_box("Error", f"An error occurred: {str(e)}", QMessageBox.Icon.Critical)

    def show_random_game_popup(self, game_name):
        """Show a pop-up window with the selected random game."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Random Game Selected")
        msg_box.setText(f"You got: {game_name}")
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        msg_box.exec()

    def lookup_game_hours(self):
        """Prompt user for a Steam App ID and show the hours played for that game."""
        # Open custom dialog to get the Steam App ID and data source choice
        dialog = GameLookupDialog(self)

        if dialog.exec() == QDialog.accepted:
            app_id, use_api = dialog.get_values()
            
            if app_id.strip():
                try:
                    # Convert to integer to validate it's a number
                    app_id = str(int(app_id.strip()))
                    
                    if use_api:
                        self.lookup_game_from_api(app_id)
                    else:
                        self.lookup_game_from_spreadsheet(app_id)
                        
                except ValueError:
                    self.show_styled_message_box("Invalid Input", "Please enter a valid numeric App ID.", QMessageBox.Warning)

    def lookup_game_from_api(self, app_id):
        """Look up game hours from Steam API."""
        try:
            import requests
            import os
            from dotenv import load_dotenv
            
            load_dotenv()
            API_KEY = os.getenv('STEAM_API_KEY')
            STEAM_ID = self.current_steam_id  # Use the current user's Steam ID
            
            if not API_KEY:
                self.show_styled_message_box("API Error", "Steam API key not found. Please check your .env file.", QMessageBox.Warning)
                return
            
            # Get owned games from Steam API
            url = f"https://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/?key={API_KEY}&steamid={STEAM_ID}&include_appinfo=true&include_played_free_games=true"
            
            response = requests.get(url)
            if response.status_code == 200:
                games_data = response.json().get('response', {}).get('games', [])
                
                # Search for the specific games
                for game in games_data:
                    if str(game.get('appid', '')) == app_id:
                        game_name = game.get('name', 'Unknown Game')
                        hours_played = round(game.get('playtime_forever', 0) / 60, 2)
                        self.show_game_hours_popup(game_name, app_id, f"{hours_played} (from API)")
                        return
                
                # Game not found
                self.show_game_not_found_popup(app_id)
            else:
                self.show_styled_message_box("API Error", f"Failed to fetch data from Steam API. Status code: {response.status_code}", QMessageBox.Warning)
                
        except ImportError:
            self.show_styled_message_box("Missing Dependency", "The 'requests' library is required for API calls.", QMessageBox.Warning)
        except Exception as e:
            self.show_styled_message_box("Error", f"An error occurred while fetching from API: {str(e)}", QMessageBox.Critical)

    def lookup_game_from_spreadsheet(self, app_id):
        """Look up game hours from spreadsheet."""
        try:
            # Ensure user directory exists before accessing spreadsheet
            self.ensure_user_directory()
            # Search for the game in the spreadsheet
            spreadsheet_path = self.get_user_spreadsheet_path()
            workbook = openpyxl.load_workbook(spreadsheet_path)
            
            if 'Steam Games Playtime' in workbook.sheetnames:
                sheet = workbook['Steam Games Playtime']
                game_found = False
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3:
                        game_name = row[0] if row[0] else "Unknown Game"  # Game Name is column A
                        sheet_app_id = str(row[1]) if row[1] else ""     # App ID is column B
                        hours_played = row[2] if row[2] else 0           # Hours is column C
                        
                        if sheet_app_id == app_id:
                            game_found = True
                            self.show_game_hours_popup(game_name, app_id, f"{hours_played} (from spreadsheet)")
                            break
                
                if not game_found:
                    self.show_game_not_found_popup(app_id)
            else:
                self.show_styled_message_box("Error", "Steam Games Playtime sheet not found in spreadsheet.", QMessageBox.Warning)
                
        except FileNotFoundError:
            self.show_styled_message_box("File Error", "Spreadsheet file not found.", QMessageBox.Warning)
        except Exception as e:
            self.show_styled_message_box("Error", f"An error occurred: {str(e)}", QMessageBox.Critical)

    def import_costs_from_csv(self):
        """Import game costs from a CSV file."""
        # Open file dialog to select CSV file
        file_dialog = QFileDialog()
        csv_file, _ = file_dialog.getOpenFileName(
            self,
            "Select CSV File",
            "ExcelFiles/",
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not csv_file:
            print("No file selected.")
            return  # User cancelled

        print(f"Selected CSV file: {csv_file}")
        import os
        if not os.path.isfile(csv_file):
            self.show_styled_message_box("File Error", f"Selected file does not exist: {csv_file}", QMessageBox.Icon.Critical)
            return

        # Create CSV importer and run import
        importer = SteamCSVImporter(parent_window=self)
        
        # Ensure user directory exists before setting spreadsheet path
        self.ensure_user_directory()
        # Set the user-specific spreadsheet path
        importer.spreadsheet_path = self.get_user_spreadsheet_path()
        
        try:
            # Patch: Style QInputDialog drop-down text to white
            from PyQt6.QtWidgets import QInputDialog
            QInputDialog.setStyleSheet(QInputDialog(), """
                QInputDialog {
                    background-color: #2b2b2b;
                    color: white;
                }
                QLabel {
                    color: white;
                }
                QComboBox {
                    background-color: #404040;
                    color: white;
                    border: 1px solid #666;
                }
                QComboBox QAbstractItemView {
                    background-color: #404040;
                    color: white;
                }
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    margin: 2px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            success, result = importer.import_from_file(csv_file)
        except Exception as e:
            self.show_styled_message_box("Import Error", f"Exception during import: {str(e)}", QMessageBox.Icon.Critical)
            return

        if success:
            # Show success message
            stats = result
            message = f"CSV import completed!\n\nGames processed: {stats['games_processed']}\nNew games added: {stats['games_added']}\nGames skipped: {stats['games_skipped']}"
            self.show_success_notification("CSV Import Complete", message)
        else:
            # Show error message
            self.show_styled_message_box("Import Error", result, QMessageBox.Icon.Critical)

    def search_game_stats(self):
        """Search for game stats by typing the game name."""
        dialog = QDialog(self)
        dialog.setWindowTitle('Search Game Stats')
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #ccc;
                padding: 5px;
                margin: 5px 0;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
            self.show_styled_message_box("Update Failed", f"Failed to update Steam data: {str(e)}", QMessageBox.Icon.Critical)
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        layout = QVBoxLayout()
        instruction_label = QLabel('Enter game name to search:')
        layout.addWidget(instruction_label)
        
        game_input = CustomTextBox()
        layout.addWidget(game_input)
        
        button_layout = QHBoxLayout()
        ok_button = QPushButton('Search')
        cancel_button = QPushButton('Cancel')
        
        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        
        dialog.setLayout(layout)
        
        if dialog.exec() != QDialog.Accepted:
            return
            
        game_name = game_input.text().strip()
        if not game_name:
            return
        
        try:
            # Ensure user directory exists before accessing spreadsheet
            self.ensure_user_directory()
            spreadsheet_path = self.get_user_spreadsheet_path()
            workbook = openpyxl.load_workbook(spreadsheet_path)
            
            if 'Steam Games Playtime' in workbook.sheetnames:
                sheet = workbook['Steam Games Playtime']
                found_games = []
                search_lower = game_name.lower()
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 6 and row[0]:
                        game_row_name = str(row[0]).strip()
                        
                        if (search_lower == game_row_name.lower() or 
                            search_lower in game_row_name.lower() or 
                            game_row_name.lower() in search_lower):
                            
                            # Calculate similarity score for sorting
                            similarity_score = calculate_similarity_score(search_lower, game_row_name.lower())
                            
                            found_games.append({
                                'name': game_row_name,
                                'app_id': row[1] if row[1] else 'N/A',
                                'hours': row[2] if row[2] else 0,
                                'cost': row[3] if row[3] else 0,
                                'date': row[4] if row[4] else 'N/A',
                                'method': row[5] if row[5] else 'N/A',
                                'similarity': similarity_score
                            })
                
                # Sort by similarity score (highest first - best matches first)
                found_games.sort(key=lambda x: x['similarity'], reverse=True)
                
                if found_games:
                    if len(found_games) == 1:
                        game = found_games[0]
                        message = f"Game: {game['name']}\n"
                        message += f"App ID: {game['app_id']}\n"
                        message += f"Hours Played: {game['hours']}\n"
                        message += f"Purchase Cost: ${game['cost']:.2f}\n"
                        message += f"Purchase Date: {game['date']}\n"
                        message += f"Purchase Method: {game['method']}"
                        
                        self.show_styled_message_box("Game Stats", message, QMessageBox.Information)
                    else:
                        self.show_multiple_game_results(found_games, game_name)
                else:
                    self.show_styled_message_box("No Results", f"No games found matching '{game_name}'.", QMessageBox.Information)
                    
            else:
                self.show_styled_message_box("Error", "Steam Games Playtime sheet not found in spreadsheet.", QMessageBox.Warning)
                
        except FileNotFoundError:
            self.show_styled_message_box("File Error", "Spreadsheet file not found.", QMessageBox.Warning)
        except Exception as e:
            self.show_styled_message_box("Error", f"An error occurred while searching: {str(e)}", QMessageBox.Critical)

    def show_multiple_game_results(self, found_games, search_term):
        """Show multiple game results in a selection dialog."""
        # Create custom dialog for multiple results
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Multiple Results for '{search_term}'")
        dialog.setMinimumSize(600, 400)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QTextEdit {
                background-color: #404040;
                color: white;
                border: 1px solid #666;
                padding: 10px;
                font-family: 'Courier New', monospace;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Title
        title_label = QLabel(f"Found {len(found_games)} games matching '{search_term}':")
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(title_label)
        
        # Results text area
        results_text = QTextEdit()
        results_text.setReadOnly(True)
        
        # Format results
        text_content = ""
        for i, game in enumerate(found_games, 1):
            text_content += f"{i}. {game['name']}"
            # Show similarity score for debugging/transparency
            if 'similarity' in game:
                text_content += f" (Match: {game['similarity']:.1f})"
            text_content += "\n"
            text_content += f"   App ID: {game['app_id']}\n"
            text_content += f"   Hours: {game['hours']}\n"
            text_content += f"   Cost: ${game['cost']:.2f}\n"
            text_content += f"   Date: {game['date']}\n"
            text_content += f"   Method: {game['method']}\n\n"
        
        results_text.setPlainText(text_content)
        layout.addWidget(results_text)
        
        # Close button
        close_button = QPushButton('Close')
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)
        
        dialog.setLayout(layout)
        dialog.exec()

    def show_success_notification(self, title, details):
        """Show a success notification popup."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(details)
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # Auto-close the dialog after 3 seconds
        QTimer.singleShot(3000, msg_box.close)
        msg_box.exec()

    def show_styled_message_box(self, title, message, icon_type):
        """Show a styled message box with white text."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.setIcon(icon_type)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        msg_box.exec()

    def show_game_hours_popup(self, game_name, app_id, hours_played):
        """Show a pop-up window with the game's hours played."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Game Hours Found")
        msg_box.setText(f"Game: {game_name}\nApp ID: {app_id}\nHours Played: {hours_played}")
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        msg_box.exec()

    def show_game_not_found_popup(self, app_id):
        """Show a styled dark pop-up when the game is not found."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Game Not Found")
        msg_box.setText(f"No game found with App ID: {app_id}\n\nThis could mean:\n- You don't own this game\n- The game wasn't included in your last data update\n- The App ID is incorrect")
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        msg_box.exec()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
