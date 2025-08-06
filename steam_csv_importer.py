from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox, QTextEdit, QProgressDialog, QApplication, QInputDialog
from PyQt6.QtGui import QFont
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import csv
import re
import openpyxl
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox, QTextEdit, QProgressDialog, QApplication, QInputDialog
from PyQt6.QtGui import QFont
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from SteamAPI_Caller import get_bundle_prices
from individual_price_dialog import IndividualPriceDialog
from game_search import calculate_similarity_score, find_best_matches, EDITION_SUFFIXES, normalize_game_name, roman_to_int, int_to_roman, normalize_numbers_in_title, extract_cost_from_string


class PriceBreakdownDialog(QDialog):
    def __init__(self, games_with_prices, total_cost, purchase_type, parent=None, 
                 steam_prices=None, total_steam_value=None, game_app_ids=None):
        super().__init__(parent)
        self.setWindowTitle(f'{purchase_type} - Price Breakdown')
        self.setMinimumSize(700, 500)  # Made wider to accommodate extra columns
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QTextEdit {
                background-color: #404040;
                color: white;
                border: 1px solid #666;
                padding: 10px;
                font-family: 'Courier New', monospace;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Title
        title_label = QLabel(f'{purchase_type} Price Breakdown')
        title_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title_label)
        
        # Summary
        summary_label = QLabel(f'Total Cost: ${total_cost:.2f} | Games: {len(games_with_prices)}')
        summary_label.setFont(QFont("Arial", 11))
        layout.addWidget(summary_label)
        
        # Game list with prices
        breakdown_text = QTextEdit()
        breakdown_text.setReadOnly(True)
        
        # Format the breakdown with enhanced information
        text_content = ""
        calculated_total = 0.0
        
        # Check if we have Steam pricing information
        has_steam_info = steam_prices and total_steam_value and game_app_ids
        
        if has_steam_info:
            # Enhanced format with Steam prices and percentages
            text_content += f"{'Game Name':<40} {'Bundle Cost':<12} {'Steam Price':<12} {'Bundle %':<10}\n"
            text_content += f"{'-'*40} {'-'*12} {'-'*12} {'-'*10}\n"
            
            for game_name, price in games_with_prices.items():
                app_id = game_app_ids.get(game_name)
                steam_price = steam_prices.get(app_id, 0.0) if app_id else 0.0
                bundle_percentage = (price / total_cost * 100) if total_cost > 0 else 0.0
                
                text_content += f"{game_name:<40} ${price:>10.2f} ${steam_price:>10.2f} {bundle_percentage:>8.1f}%\n"
                calculated_total += price
            
            text_content += f"\n{'='*76}\n"
            text_content += f"{'BUNDLE TOTAL':<40} ${calculated_total:>10.2f} ${total_steam_value:>10.2f} {'100.0%':>8}\n"
            text_content += f"{'ORIGINAL TOTAL':<40} ${total_cost:>10.2f}\n"
            
            if abs(calculated_total - total_cost) > 0.01:
                difference = calculated_total - total_cost
                text_content += f"{'DIFFERENCE':<40} ${difference:>10.2f}\n"
            
            # Add summary information
            savings = total_steam_value - total_cost
            if savings > 0:
                savings_percentage = (savings / total_steam_value * 100) if total_steam_value > 0 else 0.0
                text_content += f"\n{'BUNDLE SAVINGS':<40} ${savings:>10.2f} ({savings_percentage:.1f}% off)\n"
        else:
            # Standard format for non-weighted purchases
            for game_name, price in games_with_prices.items():
                text_content += f"{game_name:<50} ${price:>8.2f}\n"
                calculated_total += price
            
            text_content += f"\n{'='*60}\n"
            text_content += f"{'TOTAL':<50} ${calculated_total:>8.2f}\n"
            text_content += f"{'ORIGINAL TOTAL':<50} ${total_cost:>8.2f}\n"
            
            if abs(calculated_total - total_cost) > 0.01:
                difference = calculated_total - total_cost
                text_content += f"{'DIFFERENCE':<50} ${difference:>8.2f}\n"
        
        breakdown_text.setPlainText(text_content)
        layout.addWidget(breakdown_text)
        
        # Buttons
        button_layout = QHBoxLayout()
        proceed_button = QPushButton('Proceed with These Prices')
        cancel_button = QPushButton('Cancel')
        
        proceed_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(proceed_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)


class GameIdInputDialog(QDialog):
    def __init__(self, game_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Game Not Found - Enter App ID')
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #ccc;
                padding: 8px;
                margin: 5px 0;
                font-family: monospace;
            }
            QLineEdit:focus {
                border: 2px solid #4CAF50;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # Create layout
        layout = QVBoxLayout()
        
        # Add game name label
        self.game_label = QLabel(f'Game not found in library: "{game_name}"')
        layout.addWidget(self.game_label)
        
        # Add instruction
        self.instruction_label = QLabel('Please enter the Steam App ID(s) for this game:')
        layout.addWidget(self.instruction_label)
        
        # Add detailed instruction for multiple IDs
        self.multi_instruction_label = QLabel('For multiple games/DLCs, separate App IDs with commas (e.g., 271590, 271591, 271592)')
        self.multi_instruction_label.setStyleSheet("font-size: 10px; color: #cccccc; margin-bottom: 5px;")
        layout.addWidget(self.multi_instruction_label)
        
        # Add input field
        self.input_field = QLineEdit()
        self.input_field.setPlaceholderText('e.g., 271590 or 271590, 271591, 271592')
        layout.addWidget(self.input_field)
        
        # Add buttons
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton('OK')
        self.skip_button = QPushButton('Skip Game')
        self.cancel_button = QPushButton('Cancel Import')
        
        self.ok_button.clicked.connect(self.accept)
        self.skip_button.clicked.connect(self.skip)
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.skip_button)
        button_layout.addWidget(self.cancel_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # Set focus on input field
        self.input_field.setFocus()
    
    def skip(self):
        self.done(2)  # Custom return code for skip
    
    def get_app_id(self):
        """Get the first App ID (for backward compatibility)."""
        app_ids = self.get_app_ids()
        return app_ids[0] if app_ids else ""
    
    def get_app_ids(self):
        """Get all App IDs as a list."""
        text = self.input_field.text().strip()
        if not text:
            return []
        
        # Split by commas and clean up each ID
        app_ids = []
        for app_id in text.split(','):
            app_id = app_id.strip()
            if app_id and app_id.isdigit():
                app_ids.append(app_id)
        
        return app_ids
    
    def has_multiple_app_ids(self):
        """Check if multiple App IDs were entered."""
        return len(self.get_app_ids()) > 1


class BundleTypeDialog(QDialog):
    def __init__(self, bundle_games, total_cost, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Bundle Purchase Type')
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        # Create layout
        layout = QVBoxLayout()
        
        # Add title
        title_label = QLabel('Bundle Detected!')
        title_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title_label)
        
        # Show bundle info
        info_label = QLabel(f'Found {len(bundle_games)} games with total cost: ${total_cost:.2f}')
        layout.addWidget(info_label)
        
        # Show games list
        games_label = QLabel('Games:')
        layout.addWidget(games_label)
        
        games_text = '\n'.join([f'• {game}' for game in bundle_games[:5]])  # Show first 5
        if len(bundle_games) > 5:
            games_text += f'\n... and {len(bundle_games) - 5} more'
        
        games_list_label = QLabel(games_text)
        games_list_label.setStyleSheet("margin-left: 20px; font-family: monospace;")
        layout.addWidget(games_list_label)
        
        # Add question
        question_label = QLabel('\nHow should this be handled?')
        question_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(question_label)
        
        # Add explanation
        explanation = QLabel(
            "• Multi Purchase: Each game was bought separately (enter individual prices)\n"
            "• Weighted Purchase: Distribute cost based on individual Steam prices"
        )
        explanation.setStyleSheet("margin: 10px; font-size: 10px;")
        layout.addWidget(explanation)
        
        # Add buttons
        button_layout = QHBoxLayout()
        self.multi_button = QPushButton('Multi Purchase')
        self.weighted_button = QPushButton('Weighted Purchase')
        self.dlc_bundle_button = QPushButton('Base Game + DLC (Combine)')
        self.cancel_button = QPushButton('Cancel')

        self.multi_button.clicked.connect(lambda: self.done(1))
        self.weighted_button.clicked.connect(lambda: self.done(3))
        self.dlc_bundle_button.clicked.connect(lambda: self.done(5))
        self.cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(self.multi_button)
        button_layout.addWidget(self.weighted_button)
        button_layout.addWidget(self.dlc_bundle_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        self.setLayout(layout)


class ImportProgressDialog(QDialog):
    def __init__(self, total_items, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Importing Steam Games...')
        self.setMinimumSize(500, 200)
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: white;
            }
            QLabel {
                color: white;
                font-size: 12px;
            }
            QProgressBar {
                border: 2px solid #555;
                border-radius: 5px;
                text-align: center;
                background-color: #404040;
                color: white;
                font-size: 12px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)
        
        layout = QVBoxLayout()
        
        # Title label
        self.title_label = QLabel('Importing Steam game data from CSV...')
        self.title_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(self.title_label)
        
        # Current operation label
        self.status_label = QLabel('Initializing...')
        layout.addWidget(self.status_label)
        
        # Progress bar
        from PyQt6.QtWidgets import QProgressBar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, total_items)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # Stats label
        self.stats_label = QLabel('Games processed: 0 | Added: 0 | Skipped: 0')
        layout.addWidget(self.stats_label)
        
        self.setLayout(layout)
        
        # Make dialog non-closable during import
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.CustomizeWindowHint | Qt.WindowType.WindowTitleHint)
    
    def update_progress(self, current, status_text, processed=0, added=0, skipped=0):
        """Update the progress dialog with current status."""
        self.progress_bar.setValue(current)
        self.status_label.setText(status_text)
        self.stats_label.setText(f'Games processed: {processed} | Added: {added} | Skipped: {skipped}')
        QApplication.processEvents()  # Keep UI responsive
    
    def set_status(self, status_text):
        """Update just the status text."""
        self.status_label.setText(status_text)
        QApplication.processEvents()


class SteamCSVImporter:
    def __init__(self, parent_window=None):
        self.parent = parent_window
        self.spreadsheet_path = 'ExcelFiles/steam_games_playtime.xlsx'
        # Cache for Steam API calls to avoid redundant requests
        self.steam_price_cache = {}
        self.app_id_cache = {}
    
    def import_from_file(self, csv_file_path):
        """Import game costs from a CSV file."""
        try:
            # Load existing spreadsheet
            workbook = openpyxl.load_workbook(self.spreadsheet_path)
            
            # Get or create the Steam Games Playtime sheet
            if 'Steam Games Playtime' not in workbook.sheetnames:
                if self.parent:
                    self.parent.show_styled_message_box("Error", "Steam Games Playtime sheet not found in spreadsheet.", QMessageBox.Icon.Warning)
                return False, "Steam Games Playtime sheet not found"
            
            sheet = workbook['Steam Games Playtime']
            
            # Create a dictionary of existing games {app_id: row_number}
            existing_games = {}
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) >= 2 and row[1]:  # If app_id exists (App ID is column B)
                    existing_games[str(row[1])] = row_num
            
            # Process CSV file
            games_processed = 0
            games_skipped = 0
            games_added = 0
            
            # First pass: collect all purchase data to detect bundles
            purchase_data = self._parse_csv_file(csv_file_path)
            
            if not purchase_data:
                return False, "No valid purchase data found in CSV file"
            
            # Create and show progress dialog
            progress_dialog = ImportProgressDialog(len(purchase_data), self.parent)
            progress_dialog.show()
            progress_dialog.set_status("Starting import process...")
            
            current_item = 0
            
            # Second pass: prompt user for bundle handling and process games
            for purchase in purchase_data:
                current_item += 1
                
                # Update progress for this purchase
                if len(purchase['bundle_games']) > 1:
                    progress_dialog.update_progress(
                        current_item - 1, 
                        f"Processing bundle: {', '.join(purchase['bundle_games'][:2])}{'...' if len(purchase['bundle_games']) > 2 else ''}",
                        games_processed, games_added, games_skipped
                    )
                else:
                    progress_dialog.update_progress(
                        current_item - 1, 
                        f"Processing game: {purchase['bundle_games'][0]}",
                        games_processed, games_added, games_skipped
                    )
                
                if len(purchase['bundle_games']) > 1:
                    # Bundle logic
                    bundle_games = purchase['bundle_games']
                    total_cost = purchase['cost']
                    dialog = BundleTypeDialog(bundle_games, total_cost, self.parent)
                    result = dialog.exec()
                    if result == 1:
                        # ...existing code for multi purchase...
                        individual_price_dialog = IndividualPriceDialog(bundle_games, total_cost, self.parent)
                        price_result = individual_price_dialog.exec()
                        if price_result == QDialog.DialogCode.Accepted:
                            individual_prices = individual_price_dialog.get_individual_prices()
                            breakdown_dialog = PriceBreakdownDialog(
                                individual_prices, total_cost, "Multi Purchase", self.parent
                            )
                            breakdown_result = breakdown_dialog.exec()
                            if breakdown_result == QDialog.DialogCode.Accepted:
                                for game_name in bundle_games:
                                    individual_cost = individual_prices.get(game_name, 0.0)
                                    result = self._process_single_game(
                                        game_name, individual_cost, purchase['date'], "Steam",
                                        existing_games, sheet, games_processed, games_added, games_skipped
                                    )
                                    games_processed += 1
                                    if result == 'added':
                                        games_added += 1
                                    elif result == 'skipped':
                                        print(f"Game skipped: {game_name} (multi purchase)")
                                        games_skipped += 1
                            else:
                                continue
                        # No further action needed; continue to next purchase
                    elif result == 3:
                        # ...existing code for weighted purchase...
                        progress_dialog.set_status("Fetching Steam prices for weighted distribution...")
                        calc_result = self._calculate_weighted_costs(bundle_games, total_cost, sheet)
                        if calc_result and len(calc_result) == 4:
                            weighted_costs, game_app_ids, steam_prices, total_steam_value = calc_result
                            breakdown_dialog = PriceBreakdownDialog(
                                weighted_costs, total_cost, "Weighted Purchase", self.parent,
                                steam_prices=steam_prices, total_steam_value=total_steam_value, game_app_ids=game_app_ids
                            )
                            breakdown_result = breakdown_dialog.exec()
                            if breakdown_result == QDialog.DialogCode.Accepted:
                                for game_name, weighted_cost in weighted_costs.items():
                                    cached_app_id = game_app_ids.get(game_name)
                                    result = self._process_single_game_with_app_id(
                                        game_name, weighted_cost, purchase['date'], "Steam", cached_app_id,
                                        existing_games, sheet, games_processed, games_added, games_skipped
                                    )
                                    games_processed += 1
                                    if result == 'added':
                                        games_added += 1
                                    elif result == 'skipped':
                                        print(f"Game skipped: {game_name} (weighted purchase)")
                                        games_skipped += 1
                            else:
                                continue
                        else:
                            if self.parent:
                                self.parent.show_styled_message_box(
                                    "Warning",
                                    "Could not fetch Steam prices. Falling back to equal cost split.",
                                    QMessageBox.Icon.Warning
                                )
                            cost_per_game = total_cost / len(bundle_games)
                            equal_split_costs = {game: cost_per_game for game in bundle_games}
                            breakdown_dialog = PriceBreakdownDialog(
                                equal_split_costs, total_cost, "Equal Split (Fallback)", self.parent
                            )
                            breakdown_result = breakdown_dialog.exec()
                            if breakdown_result == QDialog.DialogCode.Accepted:
                                for game_name in bundle_games:
                                    result = self._process_single_game(
                                        game_name, cost_per_game, purchase['date'], "Steam",
                                        existing_games, sheet, games_processed, games_added, games_skipped
                                    )
                                    games_processed += 1
                                    if result == 'added':
                                        games_added += 1
                                    elif result == 'skipped':
                                        print(f"Game skipped: {game_name} (equal split fallback)")
                                        games_skipped += 1
                            else:
                                continue
                    elif result == 5:
                        # New: Treat bundle as base game + DLC, combine price, set to one App ID
                        # Prompt user to select the base game
                        from PyQt6.QtWidgets import QInputDialog
                        # Styled QInputDialog for base game selection
                        base_game_dialog = QInputDialog(self.parent)
                        base_game_dialog.setWindowTitle("Select Base Game")
                        base_game_dialog.setLabelText("Select the base game for this bundle:")
                        base_game_dialog.setComboBoxItems(bundle_games)
                        base_game_dialog.setStyleSheet("""
                            QInputDialog {
                                background-color: #2b2b2b;
                                color: white;
                            }
                            QLabel {
                                color: white;
                            }
                            QComboBox {
                                background-color: #404040;
                                color: white;
                                border: 1px solid #666;
                            }
                            QComboBox QAbstractItemView {
                                background-color: #404040;
                                color: white;
                            }
                            QPushButton {
                                background-color: #4CAF50;
                                color: white;
                                border: none;
                                padding: 8px 16px;
                                margin: 5px 2px;
                                min-width: 80px;
                            }
                            QPushButton:hover {
                                background-color: #45a049;
                            }
                        """)
                        ok = base_game_dialog.exec()
                        base_game = base_game_dialog.textValue()
                        if ok == QDialog.DialogCode.Accepted:
                            # Find App ID for base game
                            base_app_id = self._find_game_in_library(base_game, sheet)
                            if base_app_id:
                                row_num = existing_games.get(base_app_id)
                                if row_num:
                                    # Add total bundle cost to base game's cost
                                    prev_cost = sheet.cell(row=row_num, column=4).value or 0
                                    sheet.cell(row=row_num, column=4, value=prev_cost + total_cost)
                                    # Optionally, add DLC rows for other games
                                    for dlc_game in bundle_games:
                                        if dlc_game != base_game:
                                            new_row = sheet.max_row + 1
                                            sheet.cell(row=new_row, column=1, value=dlc_game)
                                            sheet.cell(row=new_row, column=2, value="")
                                            sheet.cell(row=new_row, column=3, value=0)
                                            sheet.cell(row=new_row, column=4, value="")
                                            sheet.cell(row=new_row, column=5, value=purchase['date'])
                                            sheet.cell(row=new_row, column=6, value="Steam")
                                            sheet.cell(row=new_row, column=7, value="DLC")
                                            sheet.cell(row=new_row, column=8, value=base_app_id)
                                    games_processed += len(bundle_games)
                                    games_added += len(bundle_games)
                                    continue
                                else:
                                    if self.parent:
                                        self.parent.show_styled_message_box("Error", f"Base game not found in spreadsheet for bundle.", QMessageBox.Icon.Warning)
                                    games_skipped += len(bundle_games)
                                    continue
                            else:
                                if self.parent:
                                    self.parent.show_styled_message_box("Error", f"Base game App ID not found for bundle.", QMessageBox.Icon.Warning)
                                games_skipped += len(bundle_games)
                                continue
                        else:
                            # User cancelled base game selection
                            games_skipped += len(bundle_games)
                            continue
                    else:
                        games_skipped += len(bundle_games)
                        continue
                else:
                    # Single game logic
                    game_name = purchase['bundle_games'][0]
                    # Only show DLC flag dialog if game does NOT have an App ID in the spreadsheet
                    app_id = self._find_game_in_library(game_name, sheet)
                    if not app_id:
        # PyQt6 imports are now at the top of the file
                        # Create QInputDialog instance for styling
                        input_dialog = QInputDialog(self.parent)
                        input_dialog.setWindowTitle("DLC Purchase Detected")
                        input_dialog.setLabelText(f"Do you want to flag '{game_name}' as a DLC and add its cost to a base game?\n\nSelect 'Yes' to choose the base game, or 'No' to treat as a regular game.")
                        input_dialog.setComboBoxItems(["No", "Yes"])
                        input_dialog.setStyleSheet("""
                            QInputDialog {
                                background-color: #2b2b2b;
                                color: white;
                            }
                            QLabel {
                                color: white;
                            }
                            QComboBox {
                                background-color: #404040;
                                color: white;
                                border: 1px solid #666;
                            }
                            QComboBox QAbstractItemView {
                                background-color: #404040;
                                color: white;
                            }
                            QPushButton {
                                background-color: #4CAF50;
                                color: white;
                                border: none;
                                padding: 8px 16px;
                                margin: 5px 2px;
                                min-width: 80px;
                            }
                            QPushButton:hover {
                                background-color: #45a049;
                            }
                        """)
                        ok = input_dialog.exec()
                        dlc_flag = input_dialog.textValue()
                        if ok == QDialog.DialogCode.Accepted and dlc_flag == "Yes":
                            # Prompt user to select base game from spreadsheet
                            base_game_names = []
                            base_game_app_ids = []
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                if len(row) >= 2 and row[0] and row[1]:
                                    base_game_names.append(str(row[0]))
                                    base_game_app_ids.append(str(row[1]))
                            # Styled QInputDialog for getItem
                            from PyQt6.QtWidgets import QInputDialog
                            QInputDialog.setStyleSheet(QInputDialog(), """
                                QInputDialog {
                                    background-color: #2b2b2b;
                                    color: white;
                                }
                                QLabel {
                                    color: white;
                                }
                                QComboBox {
                                    background-color: #404040;
                                    color: white;
                                    border: 1px solid #666;
                                }
                                QComboBox QAbstractItemView {
                                    background-color: #404040;
                                    color: white;
                                }
                                QPushButton {
                                    background-color: #4CAF50;
                                    color: white;
                                    border: none;
                                    padding: 8px 16px;
                                    margin: 5px 2px;
                                    min-width: 80px;
                                }
                                QPushButton:hover {
                                    background-color: #45a049;
                                }
                            """)
                            base_game, ok2 = QInputDialog.getItem(
                                self.parent,
                                "Select Base Game",
                                f"Select the base game for DLC '{game_name}':",
                                base_game_names,
                                0,
                                False
                            )
                            if ok2:
                                # Find base game's row
                                base_app_id = None
                                for i, name in enumerate(base_game_names):
                                    if name == base_game:
                                        base_app_id = base_game_app_ids[i]
                                        break
                                if base_app_id and base_app_id in existing_games:
                                    row_num = existing_games[base_app_id]
                                    # Add DLC cost to base game's cost (column D)
                                    prev_cost = sheet.cell(row=row_num, column=4).value or 0
                                    sheet.cell(row=row_num, column=4, value=prev_cost + purchase['cost'])
                                    # Optionally, add DLC as a separate row
                                    new_row = sheet.max_row + 1
                                    sheet.cell(row=new_row, column=1, value=game_name)  # DLC Name
                                    sheet.cell(row=new_row, column=2, value="")       # DLC has no App ID
                                    sheet.cell(row=new_row, column=3, value=0)
                                    sheet.cell(row=new_row, column=4, value=purchase['cost'])
                                    sheet.cell(row=new_row, column=5, value=purchase['date'])
                                    sheet.cell(row=new_row, column=6, value="Steam")
                                    sheet.cell(row=new_row, column=7, value="DLC")
                                    sheet.cell(row=new_row, column=8, value=base_app_id)  # Base Game App ID
                                    games_processed += 1
                                    games_added += 1
                                    continue
                                else:
                                    # Could not find base game
                                    if self.parent:
                                        self.parent.show_styled_message_box("Error", f"Base game not found for DLC '{game_name}'. Skipping.", QMessageBox.Icon.Warning)
                                    games_skipped += 1
                                    continue
                            else:
                                # User cancelled base game selection
                                games_skipped += 1
                                continue
                        result = self._process_single_game(
                            game_name, purchase['cost'], purchase['date'], "Steam",
                            existing_games, sheet, games_processed, games_added, games_skipped
                        )
                        games_processed += 1
                        if result == 'added':
                            games_added += 1
                        elif result == 'skipped':
                            games_skipped += 1
                        elif result == 'cancelled':
                            games_skipped += 1
                    else:
                        # If game has App ID, just process as normal
                        result = self._process_single_game(
                            game_name, purchase['cost'], purchase['date'], "Steam",
                            existing_games, sheet, games_processed, games_added, games_skipped
                        )
                        games_processed += 1
                        if result == 'added':
                            games_added += 1
                        elif result == 'skipped':
                            games_skipped += 1
                        elif result == 'cancelled':
                            games_skipped += 1
            
            # Final progress update
            progress_dialog.update_progress(
                len(purchase_data),
                "Saving spreadsheet...",
                games_processed, games_added, games_skipped
            )
            
            # Save the workbook
            workbook.save(self.spreadsheet_path)
            
            # Close progress dialog
            progress_dialog.close()
            
            # Return success with stats
            stats = {
                'games_processed': games_processed,
                'games_added': games_added,
                'games_skipped': games_skipped
            }
            return True, stats
            
        except FileNotFoundError:
            return False, "CSV file or spreadsheet not found"
        except Exception as e:
            # Make sure progress dialog is closed if it exists
            print(f"Exception during import_from_file: {e}")
            try:
                progress_dialog.close()
            except:
                pass
            return False, f"An error occurred during import: {str(e)}"
    
    def _parse_csv_file(self, csv_file_path):
        """Parse CSV file and extract purchase data, handling bundle format."""
        purchase_data = []
        
        try:
            # First, try to detect the delimiter
            with open(csv_file_path, 'r', encoding='utf-8') as file:
                # Read first few lines to detect delimiter
                sample_lines = [file.readline() for _ in range(3)]
                tab_count = sum(line.count('\t') for line in sample_lines)
                comma_count = sum(line.count(',') for line in sample_lines)
                delimiter = '\t' if tab_count > comma_count else ','
            
            # Now parse with the detected delimiter
            with open(csv_file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file, delimiter=delimiter)
                next(csv_reader)  # Skip header row
                
                current_bundle = None  # Track current bundle being processed
                
                for row in csv_reader:
                    if len(row) < 2:  # Need at least game name
                        continue
                    
                    # Get data from CSV
                    purchase_date = row[0].strip() if len(row) > 0 else ""
                    game_name = row[1].strip() if len(row) > 1 else ""
                    purchase_type = row[2].strip() if len(row) > 2 else ""
                    cost_str = row[3] if len(row) > 3 else ""
                    
                    # Skip if blank game name or Steam Community Market
                    if not game_name or game_name == "Steam Community Market":
                        continue
                    
                    # Check if this is a main purchase (has date and cost) or bundle item (no date/cost)
                    cost = extract_cost_from_string(cost_str)
                    
                    if purchase_date and cost > 0:
                        # This is a main purchase entry - finalize previous bundle first
                        if current_bundle:
                            purchase_data.append(current_bundle)
                        
                        # Start new bundle
                        current_bundle = {
                            'date': purchase_date,
                            'main_game': game_name,
                            'type': purchase_type,
                            'cost': cost,
                            'bundle_games': [game_name],  # Include main game as first game in bundle
                            'processed': False
                        }
                        
                    elif current_bundle and not purchase_date and cost <= 0:
                        # This is a bundle item (no date, no cost) - add to current bundle
                        current_bundle['bundle_games'].append(game_name)
                        
                    # Note: We don't need the 'else' case since all valid purchases
                    # should either be main purchases (with date/cost) or bundle items (without date/cost)
                
                # Don't forget to add the last bundle if it exists
                if current_bundle:
                    purchase_data.append(current_bundle)
                    
        except Exception as e:
            print(f"Error parsing CSV file: {e}")
        
        return purchase_data
    
    def _process_single_game_with_app_id(self, game_name, cost, date, method, app_id, existing_games, sheet, games_processed, games_added, games_skipped):
        """Process a single game entry with a known App ID (skips the lookup to avoid double dialogs)."""
        if not app_id:
            return 'skipped'
        
        # Replace cost and date for the game
        if app_id in existing_games:
            row_num = existing_games[app_id]
            # Replace cost in column D (Purchase Cost)
            sheet.cell(row=row_num, column=4, value=cost)
            # Update purchase date in column E
            sheet.cell(row=row_num, column=5, value=date)
            # Update purchase method in column F
            sheet.cell(row=row_num, column=6, value=method)
            # Set entry type in column G (import from SteamAPI_Caller for DLC detection)
            try:
                from SteamAPI_Caller import is_dlc_by_name
                entry_type = "DLC" if is_dlc_by_name(game_name) else "Game"
                sheet.cell(row=row_num, column=7, value=entry_type)
            except ImportError:
                sheet.cell(row=row_num, column=7, value="Game")  # Default fallback
            return 'updated'
        else:
            # Add new row for this game
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=1, value=game_name)  # Game Name
            sheet.cell(row=new_row, column=2, value=app_id)     # App ID
            sheet.cell(row=new_row, column=3, value=0)          # Hours Played (default to 0)
            sheet.cell(row=new_row, column=4, value=cost)       # Purchase Cost
            sheet.cell(row=new_row, column=5, value=date)       # Purchase Date
            sheet.cell(row=new_row, column=6, value=method)     # Purchase Method
            # Set entry type in column G
            try:
                from SteamAPI_Caller import is_dlc_by_name
                entry_type = "DLC" if is_dlc_by_name(game_name) else "Game"
                sheet.cell(row=new_row, column=7, value=entry_type)
            except ImportError:
                sheet.cell(row=new_row, column=7, value="Game")  # Default fallback
            sheet.cell(row=new_row, column=8, value="")         # Base Game App ID (empty for now)
            return 'added'

    def _process_single_game(self, game_name, cost, date, method, existing_games, sheet, games_processed, games_added, games_skipped):
        """Process a single game entry and add it to the spreadsheet."""
        # Check if game exists in library
        app_id = self._find_game_in_library(game_name, sheet)
        
        if not app_id:
            # Game not found, ask user for App ID
            dialog = GameIdInputDialog(game_name, self.parent)
            id_result = dialog.exec()
            
            if id_result == QDialog.DialogCode.Accepted:
                app_ids = dialog.get_app_ids()
                if app_ids:
                    if dialog.has_multiple_app_ids():
                        # Handle multiple App IDs - split cost evenly and process each
                        cost_per_game = cost / len(app_ids)
                        results = []
                        for app_id in app_ids:
                            result = self._process_single_app_id(
                                game_name, app_id, cost_per_game, date, method, 
                                existing_games, sheet
                            )
                            results.append(result)
                        # Return the result of the first game (for counting purposes)
                        return results[0] if results else 'skipped'
                    else:
                        app_id = app_ids[0]
                else:
                    if self.parent:
                        self.parent.show_styled_message_box("Invalid App ID", f"Invalid App ID entered for {game_name}. Skipping.", QMessageBox.Icon.Warning)
                    return 'skipped'
            elif id_result == 2:  # Skip button
                return 'skipped'
            else:  # Cancel button
                return 'cancelled'
        
        # Process single App ID
        return self._process_single_app_id(game_name, app_id, cost, date, method, existing_games, sheet)
    
    def _process_single_app_id(self, game_name, app_id, cost, date, method, existing_games, sheet):
        """Process a single App ID entry."""
        
        # Replace cost and date for the game
        if app_id in existing_games:
            row_num = existing_games[app_id]
            # Replace cost in column D (Purchase Price)
            sheet.cell(row=row_num, column=4, value=cost)
            # Update purchase date in column E
            sheet.cell(row=row_num, column=5, value=date)
            # Update method to Steam for CSV imports in column F
            sheet.cell(row=row_num, column=6, value=method)
            # Set entry type in column G
            try:
                from SteamAPI_Caller import is_dlc_by_name
                entry_type = "DLC" if is_dlc_by_name(game_name) else "Game"
                sheet.cell(row=row_num, column=7, value=entry_type)
            except ImportError:
                sheet.cell(row=row_num, column=7, value="Game")  # Default fallback
            return 'processed'
        else:
            # Add new row for this game
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=1, value=game_name)  # Game Name
            sheet.cell(row=new_row, column=2, value=app_id)     # App ID
            sheet.cell(row=new_row, column=3, value=0)          # Hours Played (default to 0)
            sheet.cell(row=new_row, column=4, value=cost)       # Purchase Cost
            sheet.cell(row=new_row, column=5, value=date)       # Purchase Date
            sheet.cell(row=new_row, column=6, value=method)     # Purchase Method
            # Set entry type in column G
            try:
                from SteamAPI_Caller import is_dlc_by_name
                entry_type = "DLC" if is_dlc_by_name(game_name) else "Game"
                sheet.cell(row=new_row, column=7, value=entry_type)
            except ImportError:
                sheet.cell(row=new_row, column=7, value="Game")  # Default fallback
            sheet.cell(row=new_row, column=8, value="")         # Base Game App ID (empty for now)
            return 'added'
    
    def _calculate_weighted_costs(self, bundle_games, total_cost, sheet):
        """Calculate weighted costs for bundle games based on their Steam prices."""
        try:
            # Get App IDs for all games
            app_ids = []
            game_to_app_id = {}
            
            for game_name in bundle_games:
                app_id = self._find_game_in_library(game_name, sheet)
                if not app_id:
                    # Game not found, ask user for App ID
                    # PyQt6 imports are now at the top of the file
                    dialog = GameIdInputDialog(game_name, self.parent)
                    id_result = dialog.exec()
                    
                    if id_result == QDialog.DialogCode.Accepted:
                        app_ids_list = dialog.get_app_ids()
                        if app_ids_list:
                            # For weighted costs, use the first App ID
                            app_id = app_ids_list[0]
                            if dialog.has_multiple_app_ids():
                                # Warn user that only first App ID will be used for weighted calculation
                                if self.parent:
                                    self.parent.show_styled_message_box(
                                        "Multiple App IDs", 
                                        f"Multiple App IDs entered for '{game_name}'. Using first App ID ({app_id}) for weighted cost calculation.",
                                        QMessageBox.Icon.Information
                                    )
                        else:
                            print(f"Invalid App ID entered for {game_name}. Using fallback.")
                            return None, None, None, None
                    elif id_result == 2:  # Skip button
                        print(f"Skipped App ID entry for {game_name}. Using fallback.")
                        return None, None, None, None
                    else:  # Cancel button
                        print(f"Cancelled App ID entry for {game_name}. Using fallback.")
                        return None, None, None, None
                
                if app_id:
                    app_ids.append(app_id)
                    game_to_app_id[game_name] = app_id
                    print(f"Found App ID for {game_name}: {app_id}")
                else:
                    print(f"Warning: Could not get App ID for {game_name}")
                    return None, None, None, None
            
            if not app_ids:
                print("Error: No App IDs found for bundle games")
                return None, None, None, None
            
            # Get Steam prices for all games
            steam_prices, total_steam_value = get_bundle_prices(app_ids)
            
            if total_steam_value <= 0:
                print("Error: Total Steam value is zero or negative")
                return None, None, None, None
            
            # Calculate weighted costs using original prices for weighting, bundle cost for distribution
            weighted_costs = {}
            
            print(f"\nCalculating weighted costs:")
            print(f"Bundle actual cost: ${total_cost:.2f}")
            print(f"Total Steam value (original prices): ${total_steam_value:.2f}")
            
            for game_name in bundle_games:
                app_id = game_to_app_id[game_name]
                steam_price = steam_prices[app_id]
                
                # Calculate weight as percentage of total Steam value
                weight = steam_price / total_steam_value if total_steam_value > 0 else 0
                # Apply weight to the actual bundle cost (not Steam price)
                weighted_cost = total_cost * weight
                
                weighted_costs[game_name] = weighted_cost
                
                print(f"  {game_name}: Steam ${steam_price:.2f} ({weight*100:.1f}% of Steam total) -> ${weighted_cost:.2f} of bundle")
            
            return weighted_costs, game_to_app_id, steam_prices, total_steam_value
            
        except Exception as e:
            print(f"Error calculating weighted costs: {e}")
            return None, None, None, None
    
    def _find_game_in_library(self, game_name, sheet):
        """Find a game in the library by name and return its App ID."""
        # Check cache first
        if game_name in self.app_id_cache:
            return self.app_id_cache[game_name]
        
        # Clean the game name for comparison
        clean_name = game_name.lower().strip()
        
        # Remove common suffixes that might not match exactly using centralized list
        base_name, removed_suffix = normalize_game_name(clean_name)
        
        # Try exact match first (including Roman numeral variations)
        # Test both the full name and the base name (without edition suffixes)
        search_variations = normalize_numbers_in_title(clean_name)
        base_variations = normalize_numbers_in_title(base_name) if base_name != clean_name else []
        
        # Combine all search variations (full name + base name variations)
        all_search_variations = search_variations + base_variations
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0]:  # If game name exists (Game Name is column A)
                existing_name = str(row[0]).lower().strip()
                existing_variations = normalize_numbers_in_title(existing_name)
                
                # First try exact matches
                for search_var in all_search_variations:
                    for existing_var in existing_variations:
                        if search_var == existing_var:
                            app_id = str(row[1])  # App ID from column B
                            self.app_id_cache[game_name] = app_id  # Cache the result
                            return app_id
                
                # Then try partial matches for base name (if we removed a suffix)
                if removed_suffix and base_name != clean_name:
                    for base_var in base_variations:
                        for existing_var in existing_variations:
                            # Check if base name is contained in the existing game (for series matches)
                            if base_var in existing_var or existing_var in base_var:
                                # Extra validation: ensure significant overlap
                                words_base = set(base_var.split())
                                words_existing = set(existing_var.split())
                                common_words = words_base.intersection(words_existing)
                                
                                # Require at least 2 common words for partial matches
                                if len(common_words) >= 2:
                                    app_id = str(row[1])  # App ID from column B
                                    self.app_id_cache[game_name] = app_id  # Cache the result
                                    return app_id
        
        # If we removed a suffix, try to find the base game using enhanced matching
        if removed_suffix and base_name != clean_name:
            # Collect all games from the spreadsheet for edition matching
            all_games = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[0]:  # If game name exists (Game Name is column A)
                    all_games.append({
                        'name': str(row[0]),
                        'app_id': str(row[1]),
                        'game_name': str(row[0])  # Keep both for compatibility
                    })
            
            # Use the utility function to find edition matches
            edition_matches = find_best_matches(base_name, all_games, threshold=200, max_results=3)
            
            # Add exact match bonus for sorting
            for match in edition_matches:
                if base_name == match['name'].lower().strip():
                    match['exact_match'] = True
                    match['score'] += 1000  # Boost exact matches to the top
                else:
                    match['exact_match'] = False
            
            # Sort edition matches by score (exact matches will be at top due to score boost)
            if edition_matches:
                edition_matches.sort(key=lambda x: x['score'], reverse=True)
                best_edition_match = edition_matches[0]
                
                # Found potential match - ask user for confirmation
                if self._confirm_edition_match(game_name, best_edition_match['name'], removed_suffix):
                    app_id = best_edition_match['app_id']
                    self.app_id_cache[game_name] = app_id  # Cache the result
                    return app_id
        
        # If no exact match or edition match, try enhanced similarity matching
        # Collect all games from the spreadsheet
        all_games = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2 and row[0]:  # If game name exists (Game Name is column A)
                all_games.append({
                    'name': str(row[0]),
                    'app_id': str(row[1]),
                    'game_name': str(row[0])  # Keep both for compatibility
                })
        
        # Use the utility function to find best matches
        potential_matches = find_best_matches(clean_name, all_games, threshold=50, max_results=5)
        
        # If we found potential matches, evaluate them
        if potential_matches:
            best_match = potential_matches[0]
            
            # If the best match is significantly better than others, auto-select it
            # Otherwise, use similarity score for user confirmation
            if len(potential_matches) == 1 or best_match['score'] > potential_matches[1]['score'] + 100:
                # Clear winner - use it directly
                app_id = best_match['app_id']
                self.app_id_cache[game_name] = app_id  # Cache the result
                return app_id
            else:
                # Multiple similar matches - use similarity score for user confirmation
                similarity_score = calculate_similarity_score(clean_name, best_match['name'].lower().strip())
                if similarity_score >= 200:  # Good enough match threshold
                    app_id = best_match['app_id']
                    self.app_id_cache[game_name] = app_id  # Cache the result
                    return app_id
        
        # Cache negative result to avoid repeated lookups
        self.app_id_cache[game_name] = None
        return None
    
    def _confirm_edition_match(self, searched_game, found_game, removed_suffix):
        """Ask user to confirm if a game edition matches the base game."""
        # PyQt6 imports are now at the top of the file
        # If no parent window, assume it's the same game (for testing)
        if not self.parent:
            return True
        # Create confirmation dialog
        msg_box = QMessageBox(self.parent)
        msg_box.setWindowTitle("Edition Match Found")
        msg_box.setIcon(QMessageBox.Icon.Question)
        msg_box.setText(f"Found potential match for edition:")
        msg_box.setInformativeText(
            f"Searched for: '{searched_game}'\n"
            f"Found in library: '{found_game}'\n"
            f"Removed suffix: '{removed_suffix.strip()}'\n\n"
            f"Are these the same game?"
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg_box.setDefaultButton(QMessageBox.StandardButton.Yes)
        # Apply dark theme styling
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: white;
            }
            QMessageBox QLabel {
                color: white;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                margin: 5px 2px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        result = msg_box.exec()
        return result == QMessageBox.StandardButton.Yes

# ...existing code...
